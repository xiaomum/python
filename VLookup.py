import os
from openpyxl import load_workbook
from openpyxl.styles import Font,colors,Border,Side
import pandas as pd
from pandas import DataFrame

file1 = r'D:\之江\资产管理\cmdb历史数据bak\V100+A40.xlsx'
file1_sheet1 = r'V100S'
file2 = r'D:\之江\资产管理\cmdb历史数据bak\服务器NEW_1714283496.xlsx'
file2_sheet1 = r'服务器NEW'


frame1 = pd.DataFrame(pd.read_excel(file1,file1_sheet1))
frame2 = pd.DataFrame(pd.read_excel(file2,file2_sheet1))

#result = pd.merge(frame1,frame2,how='outer',left_on='序列号',right_on='序列号')
result = pd.merge(frame1,frame2,on='序列号')
#print(result.head(20))
file3 = r'D:\之江\资产管理\cmdb历史数据bak\V100S.xlsx'
file3_sheet = r'V100S'

writer = pd.ExcelWriter(file3)
result.to_excel(writer,index=False,sheet_name=file3_sheet)
#result.to_excel(writer,index=False,ignore_index=True,sheet_name=file3_sheet)
writer.close()
#使用pd.ExcelWriter对象和to_excel将合并后的DataFrame保存成excel

wb = load_workbook(file3)
ws = wb.active

ft1 = Font(name='微软雅黑',size=10,bold=True,italic=False,color='FF000000')
ft2 = Font(name='微软雅黑',size=10,bold=False,italic=False,color='FF000000')
thin = Side(border_style='thin',color='FF000000')
border = Border(left=thin,right=thin,top=thin,bottom=thin)


for row in ws['A1:I969']:
    for cell in row:
        cell.border = border

for row in ws['A1:I1']:
    for cell in row:
        cell.font = ft1

for row in ws['A2:I969']:
    for cell in row:
        cell.font = ft2

wb.save(file3)




