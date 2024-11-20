# -*- coding: utf-8 -*-
from openpyxl import load_workbook,Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from openpyxl.drawing.image import Image
import copy

path = r'D:\咪咕文档\虚拟化资源池使用统计\资源池运行情况统计\咪咕数媒-资源池运行情况统计(2021年10月份).xlsx'
wb = load_workbook(path)
wb2 = Workbook()

sheetnames = wb.sheetnames
#添加同名sheet页
for sheetname in sheetnames:
    print(sheetname)
    sheet = wb[sheetname]
    sheet2 = wb2.create_sheet(sheetname,index=None)

    #工作表sheet标签颜色
    sheet2.sheet_properties.tabColor = sheet.sheet_properties.tabColor

    #开始处理合并单元格形式为“(<CellRange C2：E3>,)，替换掉(<CellRange 和 >,)' 找到合并单元格
    wm = list(sheet.merged_cells)
    if len(wm) > 0:
        for i in range(0,len(wm)):
            cell2 = str(wm[i]).replace('(<CellRange','').replace('>,)','')
            sheet2.merge_cells(cell2)

    #设置单元格行高和列宽
    for i,row in enumerate(sheet.iter_rows()):
        sheet2.row_dimensions[i+1].height = sheet.row_dimensions[i+1].height
        for j,cell in enumerate(row):
            sheet2.column_dimensions[get_column_letter(j+1)].width = sheet.column_dimensions[get_column_letter(j+1)].width
            sheet2.cell(row=i+1,column=j+1,value=cell.value)

            #设置单元格格式,包括填充色、字体、样式、边框、对齐方式、数字格式、写保护
            source_cell = sheet.cell(i+1,j+1)
            target_cell = sheet2.cell(i+1,j+1)
            target_cell.fill = copy.copy(source_cell.fill) #填充色
            if source_cell.has_style:
               target_cell._style = copy.copy(source_cell._style)
               target_cell.font = copy.copy(source_cell.font)
               target_cell.border = copy.copy(source_cell.border)
               target_cell.fill = copy.copy(source_cell.fill)
               target_cell.number_format = copy.copy(source_cell.number_format)
               target_cell.protection = copy.copy(source_cell.protection)
               target_cell.alignment = copy.copy(source_cell.alignment)
    for image in sheet._images:
        sheet2.add_image

if 'Sheet' in wb2.sheetnames:
    del wb2['Sheet']

#复制图片到封面
img = Image(r'D:\咪咕文档\虚拟化资源池使用统计\sm.png')
ws = wb2.worksheets[0]
img.anchor = 'C14'
img.height = 400
img.width = 400
ws.add_image(img)
wb2.save(r'D:\咪咕文档\虚拟化资源池使用统计\result.xlsx')

#设置sheet页指定单元格区域填充色
def Textcolor(file_name):
    wb2 = load_workbook(file_name)
    sheetnames = wb2.sheetnames
    for sheetname in sheetnames:
        sheet2 = wb2[sheetname]
        for row in sheet2 ['A1:H4']:
            for cell in row:
                Color = ['000000', 'FFFFFF', '0000FF']
                fill = PatternFill('solid', fgColor=Color[1])
                cell.fill = fill
    wb2.save(file_name)

if __name__ == '__main__':
    file_name = r'D:\咪咕文档\虚拟化资源池使用统计\result.xlsx'
    Textcolor(file_name)






