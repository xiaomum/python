#author-zpy
from openpyxl import load_workbook
from openpyxl import workbook
from openpyxl.cell import WriteOnlyCell #单元格样式或注释
import os

from openpyxl.utils.cell import col
wb = load_workbook(r'D:\之江\资产管理\CMDB数据.xlsx')
ws = wb['物理服务器']
#for row in ws.iter_cols(min_row=1,max_col=6,max_row=39):
    #for cell in col:
        #print(cell)
for col in ws:
    for value in col:
        print(value)
