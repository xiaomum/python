#author-zpy
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
#from PIL import Image
import os
path =r'D:\咪咕文档\虚拟化资源池使用统计\test'
file_list = os.listdir(path)
for file in file_list:
    new_wb = load_workbook(path + '/' + file)
    new_ws =new_wb['封面']
wb = load_workbook(r'D:\咪咕文档\虚拟化资源池使用统计\资源池运行情况月报\咪咕数媒-资源池运行情况统计8月份数据.xlsx')
ws = wb['封面']
for row in ws.iter_rows(min_row=1,values_only=True):
    new_ws.append(row)
img = Image('C:\\Users\\zpy\\Desktop\\图片1.png')
new_ws.add_image(img)

new_ws.move_range('A1:E1',rows=2,cols=2,translate=True)
new_ws.insert_rows(2)
new_wb.save(r'D:\咪咕文档\虚拟化资源池使用统计\test\test.xlsx')
