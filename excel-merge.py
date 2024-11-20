#author_zpy
import os
from openpyxl import Workbook
from openpyxl import load_workbook

new_wb = Workbook()
new_ws1 = new_wb.active
new_ws2 = new_wb.add_sheet(1)
new_ws3 = new_wb.create_sheet(2)
new_ws1.title = '物理服务器'
new_ws2.title = '网络设备'
new_ws3.title = '机架资源'

path =r'D:\之江\资产管理\1'
filelist = os.listdir(path)
for file in filelist:
    wb = load_workbook(path + '/' + file)
    ws1 = wb['物理服务器']
    ws2 = wb['网络设备']
    ws3 = wb['机架资源']
    #file_link = '=HYPERLINK("'+file_path+'")'
    for row in ws1.iter_rows(min_row=1,values_only=True):
        new_ws1.append(row)
    for row in ws2.iter_rows(min_row=1,values_only=True):
        new_ws2.append(row)
    for row in ws3.iter_rows(min_row=1,values_only=True):
        new_ws3.append(row)

new_wb.save(r'D:\之江\资产管理\1\cmdb-bak.xlsx')