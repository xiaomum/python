#author-zpy
from openpyxl import load_workbook
from openpyxl.styles import colors
from openpyxl.styles import Font,Color

wb = load_workbook(r'D:\咪咕文档\虚拟化资源池使用统计\first.xlsx')
ws = wb ['修订记录']
a1 = ws['B3']
b1 = ws['C3']
ft = Font(color=colors.RED)
a1.font = ft
b1.font = ft
a1.font.italic = True
b1.font.itatlic = True
a1.font = Font(color=colors.RED, italic=True)
wb.save(r'D:\咪咕文档\虚拟化资源池使用统计\test\test.xlsx')

#PatternFill (填充类)：颜色等
#font(字体类)：字号、字体颜色、下划线等
#border(边框类)：设置单元格边框
#alignment(位置类)：对齐方式
#number_format(格式类)：数据格式
#protection(保护类)：写保护
#side 边框位置